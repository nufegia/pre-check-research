from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

import numpy as np
import pandas as pd


def source_path(path: str) -> Path:
    source = Path(path).expanduser().resolve()
    if not source.exists():
        raise FileNotFoundError(f"Input file not found: {source}")
    return source


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def markdown_out(path_str: str) -> Path:
    path = Path(path_str).expanduser().resolve()
    if path.suffix != ".md":
        path = path.with_suffix(".md")
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def clean_cell(value: Any) -> Any:
    if value is None:
        return np.nan
    if isinstance(value, str):
        value = value.strip()
        return np.nan if value == "" else value
    return value


def read_csv(path: Path) -> list[tuple[str, pd.DataFrame]]:
    encodings = ["utf-8-sig", "utf-8", "gb18030", "latin1"]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            return [(path.stem, pd.read_csv(path, encoding=encoding).map(clean_cell))]
        except Exception as exc:
            last_error = exc
    raise ValueError(f"Cannot read CSV: {last_error}")


def read_excel(path: Path) -> list[tuple[str, pd.DataFrame]]:
    if path.suffix.lower() == ".xlsx":
        segmented = read_excel_layout_tables(path)
        if segmented:
            return segmented
    sheets = pd.read_excel(path, sheet_name=None)
    return [(name, df.map(clean_cell)) for name, df in sheets.items()]


def read_excel_layout_tables(path: Path) -> list[tuple[str, pd.DataFrame]]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except ImportError:
        return []

    wb = load_workbook(path, data_only=True, read_only=False)
    tables: list[tuple[str, pd.DataFrame]] = []
    any_segmented = False
    for ws in wb.worksheets:
        values: dict[tuple[int, int], Any] = {}
        bordered: set[tuple[int, int]] = set()
        for row in ws.iter_rows():
            for cell in row:
                cleaned = clean_cell(cell.value)
                if not pd.isna(cleaned):
                    values[(cell.row, cell.column)] = cleaned
                border = cell.border
                if any(getattr(border, side).style for side in ("left", "right", "top", "bottom")):
                    bordered.add((cell.row, cell.column))

        for merged_range in ws.merged_cells.ranges:
            top_left = values.get((merged_range.min_row, merged_range.min_col))
            if top_left is None:
                continue
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    values.setdefault((row, col), top_left)

        if not values:
            continue

        candidates = _excel_layout_candidates(values, bordered)
        if len(candidates) <= 1:
            fallback_df = pd.read_excel(path, sheet_name=ws.title).map(clean_cell)
            tables.append((ws.title, fallback_df))
            continue

        any_segmented = True
        for idx, box in enumerate(_split_boxes_on_empty_axes(values, candidates), start=1):
            df = _dataframe_from_excel_box(values, box)
            if df.empty:
                continue
            top, left, bottom, right = box
            ref = f"{get_column_letter(left)}{top}_{get_column_letter(right)}{bottom}"
            caption = _nearby_excel_caption(values, box)
            label = f"{ws.title}__{caption}__range_{ref}" if caption else f"{ws.title}__range_{ref}"
            tables.append((label, df.map(clean_cell)))
    return tables if any_segmented else []


def _excel_layout_candidates(
    values: dict[tuple[int, int], Any], bordered: set[tuple[int, int]]
) -> list[tuple[int, int, int, int]]:
    candidates: list[tuple[int, int, int, int]] = []
    for component in _grid_components(bordered):
        box = _component_box(component)
        trimmed = _trim_box_to_values(values, box)
        if trimmed and _box_value_count(values, trimmed) >= 4:
            candidates.append(trimmed)

    for component in _grid_components(set(values)):
        box = _component_box(component)
        if _box_value_count(values, box) < 4:
            continue
        if any(_boxes_overlap(existing, box) for existing in candidates):
            continue
        candidates.append(box)

    return _dedupe_boxes(candidates)


def _grid_components(cells: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    seen: set[tuple[int, int]] = set()
    components: list[set[tuple[int, int]]] = []
    for cell in sorted(cells):
        if cell in seen:
            continue
        stack = [cell]
        seen.add(cell)
        component: set[tuple[int, int]] = set()
        while stack:
            row, col = stack.pop()
            component.add((row, col))
            for neighbor in ((row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1)):
                if neighbor in cells and neighbor not in seen:
                    seen.add(neighbor)
                    stack.append(neighbor)
        components.append(component)
    return components


def _component_box(component: set[tuple[int, int]]) -> tuple[int, int, int, int]:
    rows = [row for row, _ in component]
    cols = [col for _, col in component]
    return min(rows), min(cols), max(rows), max(cols)


def _trim_box_to_values(
    values: dict[tuple[int, int], Any], box: tuple[int, int, int, int]
) -> tuple[int, int, int, int] | None:
    top, left, bottom, right = box
    occupied = [(row, col) for row, col in values if top <= row <= bottom and left <= col <= right]
    if not occupied:
        return None
    rows = [row for row, _ in occupied]
    cols = [col for _, col in occupied]
    return min(rows), min(cols), max(rows), max(cols)


def _box_value_count(values: dict[tuple[int, int], Any], box: tuple[int, int, int, int]) -> int:
    top, left, bottom, right = box
    return sum(1 for row, col in values if top <= row <= bottom and left <= col <= right)


def _box_contains(outer: tuple[int, int, int, int], inner: tuple[int, int, int, int]) -> bool:
    return outer[0] <= inner[0] and outer[1] <= inner[1] and outer[2] >= inner[2] and outer[3] >= inner[3]


def _boxes_overlap(left: tuple[int, int, int, int], right: tuple[int, int, int, int]) -> bool:
    return not (left[2] < right[0] or right[2] < left[0] or left[3] < right[1] or right[3] < left[1])


def _dedupe_boxes(boxes: list[tuple[int, int, int, int]]) -> list[tuple[int, int, int, int]]:
    ordered = sorted(set(boxes), key=lambda box: (box[0], box[1], box[2], box[3]))
    deduped: list[tuple[int, int, int, int]] = []
    for box in ordered:
        if any(_box_contains(existing, box) for existing in deduped):
            continue
        deduped.append(box)
    return deduped


def _split_boxes_on_empty_axes(
    values: dict[tuple[int, int], Any], boxes: list[tuple[int, int, int, int]]
) -> list[tuple[int, int, int, int]]:
    split: list[tuple[int, int, int, int]] = []
    for box in boxes:
        sub_boxes = [box]
        changed = True
        while changed:
            changed = False
            next_boxes: list[tuple[int, int, int, int]] = []
            for sub_box in sub_boxes:
                parts = _split_box_once_on_empty_axis(values, sub_box)
                if len(parts) > 1:
                    changed = True
                next_boxes.extend(parts)
            sub_boxes = next_boxes
        split.extend(sub_boxes)
    return _dedupe_boxes([box for box in split if _box_value_count(values, box) >= 4])


def _split_box_once_on_empty_axis(
    values: dict[tuple[int, int], Any], box: tuple[int, int, int, int]
) -> list[tuple[int, int, int, int]]:
    top, left, bottom, right = box
    for row in range(top + 1, bottom):
        if not any((row, col) in values for col in range(left, right + 1)):
            return [(top, left, row - 1, right), (row + 1, left, bottom, right)]
    for col in range(left + 1, right):
        if not any((row, col) in values for row in range(top, bottom + 1)):
            return [(top, left, bottom, col - 1), (top, col + 1, bottom, right)]
    return [box]


def _dataframe_from_excel_box(values: dict[tuple[int, int], Any], box: tuple[int, int, int, int]) -> pd.DataFrame:
    top, left, bottom, right = box
    rows = [[values.get((row, col), np.nan) for col in range(left, right + 1)] for row in range(top, bottom + 1)]
    df = pd.DataFrame(rows).dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.shape[0] >= 2:
        first = ["" if pd.isna(item) else str(item) for item in df.iloc[0].tolist()]
        has_header = len(set(first)) == len(first) and any(not _looks_numeric(item) for item in first)
        if has_header:
            body = df.iloc[1:].reset_index(drop=True)
            body.columns = first
            df = body
    return df.reset_index(drop=True)


def _nearby_excel_caption(values: dict[tuple[int, int], Any], box: tuple[int, int, int, int]) -> str:
    top, left, _bottom, right = box
    for row in range(top - 1, max(top - 3, 0), -1):
        labels = [str(values[(row, col)]).strip() for col in range(left, right + 1) if (row, col) in values]
        short_labels = [
            label
            for label in labels
            if label and len(label) <= 60 and any(ch.isalpha() for ch in label) and not _looks_numeric(label)
        ]
        if len(short_labels) > 3:
            continue
        if short_labels:
            return safe_table_name("_".join(short_labels), 0)
    return ""


def _looks_numeric(value: Any) -> bool:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return False
    text = str(value).strip().replace(",", "").replace("%", "")
    if text == "":
        return False
    try:
        float(text)
    except ValueError:
        return False
    return True


def read_docx_tables(path: Path) -> list[tuple[str, pd.DataFrame]]:
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as zf:
        document = zf.read("word/document.xml")
    root = ElementTree.fromstring(document)
    tables: list[tuple[str, pd.DataFrame]] = []
    for idx, table in enumerate(root.findall(".//w:tbl", ns), start=1):
        rows = []
        for tr in table.findall(".//w:tr", ns):
            cells = []
            for tc in tr.findall("./w:tc", ns):
                text_parts = [node.text or "" for node in tc.findall(".//w:t", ns)]
                cells.append("".join(text_parts).strip())
            if cells:
                rows.append(cells)
        if not rows:
            continue
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        header = normalized[0]
        body = normalized[1:] if len(normalized) > 1 else []
        tables.append((f"docx_table_{idx}", pd.DataFrame(body, columns=header).map(clean_cell)))
    return tables


def read_pdf_tables(path: Path) -> list[tuple[str, pd.DataFrame]]:
    import pdfplumber

    tables: list[tuple[str, pd.DataFrame]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page_idx, page in enumerate(pdf.pages, start=1):
            for table_idx, table in enumerate(page.extract_tables() or [], start=1):
                if not table:
                    continue
                width = max(len(row or []) for row in table)
                normalized = [(row or []) + [""] * (width - len(row or [])) for row in table]
                header = [str(item or "") for item in normalized[0]]
                body = normalized[1:] if len(normalized) > 1 else []
                tables.append((f"pdf_p{page_idx}_table_{table_idx}", pd.DataFrame(body, columns=header).map(clean_cell)))
    return tables


def load_tables(path: Path) -> list[tuple[str, pd.DataFrame]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xls"}:
        return read_excel(path)
    if suffix == ".docx":
        return read_docx_tables(path)
    if suffix == ".pdf":
        return read_pdf_tables(path)
    raise ValueError(f"Unsupported file type: {suffix}")


def read_text_source(source: Path) -> str:
    suffix = source.suffix.lower()
    if suffix in {".txt", ".md", ".bib", ".ris"}:
        return source.read_text(encoding="utf-8")
    if suffix == ".docx":
        with zipfile.ZipFile(str(source)) as zf:
            doc = zf.read("word/document.xml")
        root = ElementTree.fromstring(doc)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        paragraphs = []
        for paragraph in root.findall(".//w:p", ns):
            texts = [text.text or "" for text in paragraph.findall(".//w:t", ns)]
            paragraphs.append("".join(texts))
        return "\n".join(paragraphs)
    if suffix == ".pdf":
        import pdfplumber

        texts = []
        with pdfplumber.open(str(source)) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    texts.append(text)
        return "\n".join(texts)
    return ""


def write_extracted_text(source: Path, workdir: Path) -> Path | None:
    if source.suffix.lower() not in {".pdf", ".docx"}:
        return None
    text = read_text_source(source)
    if not text:
        return None
    out_path = workdir / f"{source.stem}_extracted.txt"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(text, encoding="utf-8")
    return out_path


def safe_table_name(name: str, index: int) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in name.strip())
    return cleaned or f"table_{index}"


def extract_file(source: Path, out_dir: Path) -> dict[str, Any]:
    out_dir.mkdir(parents=True, exist_ok=True)
    outputs: list[dict[str, Any]] = []
    for idx, (name, df) in enumerate(load_tables(source), start=1):
        out_path = out_dir / f"{idx:02d}_{safe_table_name(name, idx)}.csv"
        df.to_csv(out_path, index=False)
        outputs.append(
            {
                "name": name,
                "kind": "table",
                "path": str(out_path),
                "rows": int(df.shape[0]),
                "columns": int(df.shape[1]),
            }
        )
    notes = []
    if source.suffix.lower() in {".pdf", ".docx"}:
        notes.append("Input was extracted from a document; merged cells, footnotes, and complex layouts may require manual review.")
    return {
        "source": str(source),
        "tool_id": "pcr_extract",
        "tool_name": "PCR Extract",
        "detector_runtime": "python",
        "dependency_status": "ready",
        "outputs": outputs,
        "notes": notes,
    }
