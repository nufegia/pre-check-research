from __future__ import annotations

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Side

from pcr_audit.io import (
    clean_cell,
    extract_file,
    load_tables,
    markdown_out,
    read_csv,
    read_excel,
    read_json,
    read_text_source,
    safe_table_name,
    source_path,
    write_json,
)


class TestCleanCell:
    def test_none_returns_nan(self):
        result = clean_cell(None)
        assert pd.isna(result)

    def test_empty_string_returns_nan(self):
        result = clean_cell("")
        assert pd.isna(result)

    def test_whitespace_only_returns_nan(self):
        result = clean_cell("   ")
        assert pd.isna(result)

    def test_strips_whitespace(self):
        result = clean_cell("  hello  ")
        assert result == "hello"

    def test_preserves_non_empty_string(self):
        result = clean_cell("hello")
        assert result == "hello"


class TestSourcePath:
    def test_raises_on_missing_file(self):
        with pytest.raises(FileNotFoundError):
            source_path("/nonexistent/path/file.csv")

    def test_returns_resolved_path(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("a,b\n1,2")
        result = source_path(str(f))
        assert result.exists()
        assert result.name == "test.csv"


class TestWriteReadJson:
    def test_write_and_read_roundtrip(self, tmp_path):
        data = {"key": "value", "nested": {"a": 1}}
        p = tmp_path / "test.json"
        write_json(p, data)
        assert p.exists()
        loaded = read_json(p)
        assert loaded == data

    def test_write_creates_parent_dirs(self, tmp_path):
        p = tmp_path / "deep" / "nested" / "test.json"
        write_json(p, {"x": 1})
        assert p.exists()


class TestMarkdownOut:
    def test_ensures_md_suffix(self):
        result = markdown_out("/tmp/report")
        assert result.suffix == ".md"

    def test_preserves_md_suffix(self):
        result = markdown_out("/tmp/report.md")
        assert result.suffix == ".md"


class TestReadCsv:
    def test_reads_simple_csv(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("a,b,c\n1,2,3\n4,5,6")
        tables = read_csv(f)
        assert len(tables) == 1
        name, df = tables[0]
        assert name == "test"
        assert df.shape == (2, 3)
        assert list(df.columns) == ["a", "b", "c"]

    def test_handles_utf8_bom(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_bytes(b"\xef\xbb\xbfa,b\n1,2")
        tables = read_csv(f)
        assert len(tables) == 1


class TestReadExcel:
    def test_reads_single_sheet(self, tmp_path):
        f = tmp_path / "test.xlsx"
        df = pd.DataFrame({"x": [1, 2], "y": [3, 4]})
        df.to_excel(f, index=False)
        tables = read_excel(f)
        assert len(tables) >= 1
        name, result_df = tables[0]
        assert result_df.shape == (2, 2)

    def test_splits_bordered_layout_tables(self, tmp_path):
        f = tmp_path / "layout.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Source Data"
        ws["B2"] = "Fig. 1a"
        ws["B3"] = "group"
        ws["C3"] = "value"
        ws["B4"] = "A"
        ws["C4"] = 1
        ws["B5"] = "B"
        ws["C5"] = 2
        ws["E2"] = "Fig. 1b"
        ws["E3"] = "group"
        ws["F3"] = "value"
        ws["E4"] = "A"
        ws["F4"] = 3
        ws["E5"] = "B"
        ws["F5"] = 4

        side = Side(style="thin", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for row in range(3, 6):
            for col in (2, 3, 5, 6):
                ws.cell(row=row, column=col).border = border
        wb.save(f)

        tables = read_excel(f)

        assert [name for name, _df in tables] == [
            "Source Data__Fig__1a__range_B3_C5",
            "Source Data__Fig__1b__range_E3_F5",
        ]
        assert [df.shape for _name, df in tables] == [(2, 2), (2, 2)]
        assert tables[0][1].iloc[0].to_dict() == {"group": "A", "value": 1}

    def test_preserves_unsegmented_sheets_in_mixed_workbook(self, tmp_path):
        f = tmp_path / "mixed.xlsx"
        wb = Workbook()
        layout = wb.active
        layout.title = "Layout"
        plain = wb.create_sheet("Plain")
        side = Side(style="thin", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for left, value in ((1, 10), (4, 20)):
            layout.cell(row=1, column=left, value="id").border = border
            layout.cell(row=1, column=left + 1, value="value").border = border
            layout.cell(row=2, column=left, value="x").border = border
            layout.cell(row=2, column=left + 1, value=value).border = border
            layout.cell(row=3, column=left, value="y").border = border
            layout.cell(row=3, column=left + 1, value=value + 1).border = border
        plain.append(["subject", "score"])
        plain.append(["s1", 1])
        plain.append(["s2", 2])
        wb.save(f)

        tables = read_excel(f)

        assert [name for name, _df in tables] == ["Layout__range_A1_B3", "Layout__range_D1_E3", "Plain"]
        assert tables[-1][1].shape == (2, 2)


class TestReadTextSource:
    def test_reads_txt_file(self, tmp_path):
        f = tmp_path / "test.txt"
        f.write_text("Hello world.")
        text = read_text_source(f)
        assert "Hello world." in text

    def test_reads_markdown_file(self, tmp_path):
        f = tmp_path / "test.md"
        f.write_text("# Title\n\nContent.")
        text = read_text_source(f)
        assert "Title" in text


class TestLoadTables:
    def test_loads_csv(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("a,b\n1,2")
        tables = load_tables(f)
        assert len(tables) == 1

    def test_rejects_unsupported_type(self, tmp_path):
        f = tmp_path / "test.xyz"
        f.write_text("data")
        with pytest.raises(ValueError, match="Unsupported file type"):
            load_tables(f)


class TestSafeTableName:
    def test_preserves_alphanumeric(self):
        assert safe_table_name("hello_123", 1) == "hello_123"

    def test_replaces_special_chars(self):
        result = safe_table_name("sheet (1)", 1)
        assert "(" not in result

    def test_falls_back_to_index_for_empty(self):
        result = safe_table_name("   ", 5)
        assert result == "table_5"


class TestExtractFile:
    def test_extracts_csv_to_csv(self, tmp_path):
        f = tmp_path / "test.csv"
        f.write_text("a,b\n1,2\n3,4")
        out_dir = tmp_path / "out"
        manifest = extract_file(f, out_dir)
        assert manifest["source"] == str(f)
        assert len(manifest["outputs"]) == 1
        assert manifest["outputs"][0]["kind"] == "table"
        csv_out = tmp_path / "out" / "01_test.csv"
        assert csv_out.exists()

    def test_extracts_bordered_excel_tables_to_separate_csvs(self, tmp_path):
        f = tmp_path / "layout.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet"
        side = Side(style="thin", color="000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for left, value in ((1, 10), (4, 20)):
            ws.cell(row=1, column=left, value="id").border = border
            ws.cell(row=1, column=left + 1, value="value").border = border
            ws.cell(row=2, column=left, value="x").border = border
            ws.cell(row=2, column=left + 1, value=value).border = border
            ws.cell(row=3, column=left, value="y").border = border
            ws.cell(row=3, column=left + 1, value=value + 1).border = border
        wb.save(f)

        manifest = extract_file(f, tmp_path / "out")

        assert len(manifest["outputs"]) == 2
        assert (tmp_path / "out" / "01_Sheet__range_A1_B3.csv").exists()
        assert (tmp_path / "out" / "02_Sheet__range_D1_E3.csv").exists()
